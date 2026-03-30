from __future__ import annotations

import pandas as pd
import numpy as np

from optimizador import (
    optimizar_rotacion,
    preparar_mapas_fechas,
    reordenar_columnas_reporte,
)

from config import *

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.numbers import FORMAT_DATE_XLSX14


# =====================================================
# ARMAR PAQUETE
# =====================================================

def armar_paquete_v4(
    resultado: pd.DataFrame,
    df_inventario_rotar_final: pd.DataFrame,
    df_referencia_ciudad: pd.DataFrame,
    df_traslados_costos: pd.DataFrame,
    etiqueta_modo: str,
    map_fecha: pd.Series,
    map_dias_ingreso: pd.Series
):

    placas_con_modelo = resultado["Placa"].unique()

    no_movibles = df_inventario_rotar_final[
        ~df_inventario_rotar_final["Placa"].isin(placas_con_modelo)
    ].copy()

    # =========================
    # RESUMEN INICIAL
    # =========================
    resumen_inicial = (
        df_inventario_rotar_final.groupby("UbicacionActual")
        .agg(Inicial_Total=("Placa", "count"))
        .reset_index()
        .rename(columns={"UbicacionActual": "Ciudad"})
    )

    total_inicial = len(df_inventario_rotar_final)
    resumen_inicial["Inicial_%"] = (
        resumen_inicial["Inicial_Total"] / total_inicial * 100
    ).round(2)

    # =========================
    # ESCENARIO FINAL (RANK 1)
    # =========================
    resultado_rank1 = resultado[resultado["Rank"] == 1].copy()

    placas_rank1 = resultado_rank1["Placa"].unique()

    no_movibles_rank1 = df_inventario_rotar_final[
        ~df_inventario_rotar_final["Placa"].isin(placas_rank1)
    ].copy()

    df_movidas = pd.DataFrame({
        "Placa": resultado_rank1["Placa"],
        "Ciudad_Origen": resultado_rank1["Ciudad_Origen"],
        "Ciudad_Final": resultado_rank1["Ciudad_Recomendada"],
        "Costo_Traslado": resultado_rank1["Costo_Traslado"]
    })

    df_quedan = pd.DataFrame({
        "Placa": no_movibles_rank1["Placa"],
        "Ciudad_Origen": no_movibles_rank1["UbicacionActual"],
        "Ciudad_Final": no_movibles_rank1["UbicacionActual"],
        "Costo_Traslado": 0
    })

    df_estado_final = pd.concat([df_movidas, df_quedan], ignore_index=True)
    df_estado_final["Es_Traslado"] = (
        df_estado_final["Ciudad_Final"] != df_estado_final["Ciudad_Origen"]
    )

    total_final = len(df_estado_final)

    resumen_final = (
        df_estado_final.groupby("Ciudad_Final")
        .agg(
            Final_Total=("Placa", "count"),
            Traslados_Recibidos=("Es_Traslado", "sum"),
            Costo_Traslado_Total=("Costo_Traslado", "sum")
        )
        .reset_index()
        .rename(columns={"Ciudad_Final": "Ciudad"})
    )

    resumen_final["Final_%"] = (
        resumen_final["Final_Total"] / total_final * 100
    ).round(2)

    resumen_comparado = (
        resumen_inicial.merge(resumen_final, on="Ciudad", how="outer")
        .fillna(0)
    )

    # =========================
    # TRASLADOS
    # =========================
    df_quedarse = (
        resultado[
            resultado["Ciudad_Recomendada"] == resultado["Ciudad_Origen"]
        ][["Placa", "Dias_Esperados", "Costo_Total"]]
        .rename(columns={
            "Dias_Esperados": "Dias_Esperados_Origen",
            "Costo_Total": "Costo_Quedarse"
        })
        .sort_values("Costo_Quedarse")
        .drop_duplicates(subset=["Placa"])
    )

    df_mover = (
        resultado_rank1[
            [
                "Placa",
                "Referencia",
                "Ciudad_Origen",
                "Ciudad_Recomendada",
                "Dias_Esperados",
                "Costo_Total",
                "Margen_Incremental"
            ]
        ]
        .rename(columns={
            "Dias_Esperados": "Dias_Esperados_Destino",
            "Costo_Total": "Costo_Moverme"
        })
    )

    df_traslados = df_mover.merge(df_quedarse, on="Placa", how="left")

    df_traslados["Diferencia_Costos"] = (
        df_traslados["Costo_Quedarse"] -
        df_traslados["Costo_Moverme"]
    )

    df_traslados = (
        df_traslados[
            df_traslados["Diferencia_Costos"] != 0
        ]
        .sort_values("Diferencia_Costos", ascending=False)
        .reset_index(drop=True)
    )

    # =========================
    # FECHA + DIAS
    # =========================
    resultado_out = resultado.copy()
    resultado_out["FechaEntregaMoto"] = resultado_out["Placa"].map(map_fecha)
    resultado_out["Dias_Desde_Ingreso"] = resultado_out["Placa"].map(map_dias_ingreso)

    resultado_out = reordenar_columnas_reporte(resultado_out)

    df_origen_dias = (
        resultado_out[
            resultado_out["Ciudad_Recomendada"] ==
            resultado_out["Ciudad_Origen"]
        ][["Placa", "Dias_Esperados"]]
        .rename(columns={
            "Dias_Esperados": "Dias_Esperados_Origen"
        })
        .drop_duplicates(subset=["Placa"])
    )

    resultado_out = resultado_out.merge(
        df_origen_dias,
        on="Placa",
        how="left"
    )

    resultado_out = resultado_out.rename(
        columns={"Dias_Esperados": "Dias_Esperados_Destino"}
    )

    df_traslados_out = df_traslados.copy()
    df_traslados_out["FechaEntregaMoto"] = df_traslados_out["Placa"].map(map_fecha)
    df_traslados_out["Dias_Desde_Ingreso"] = df_traslados_out["Placa"].map(map_dias_ingreso)
    df_traslados_out = reordenar_columnas_reporte(df_traslados_out)

    return {
        "resultado": resultado_out,
        "traslados": df_traslados_out,
        "no_movibles": no_movibles,
        "resumen": resumen_comparado
    }

# =====================================================
# EXPORTADOR
# =====================================================

def exportar_dual_v4(
    df_inventario_rotar_final: pd.DataFrame,
    df_referencia_ciudad: pd.DataFrame,
    df_traslados: pd.DataFrame,
    df_custodia: pd.DataFrame,
    top_n: int = 999,
    filename: str = "Reporte_Modelo_Final_Dual_V4.xlsx"
):

    map_fecha, map_dias_ingreso = preparar_mapas_fechas(
        df_inventario_rotar_final
    )

    placas_inventario = df_inventario_rotar_final["Placa"].tolist()

    resultado_estricto = optimizar_rotacion(
        placas=placas_inventario,
        df_inventario=df_inventario_rotar_final,
        df_stats=df_referencia_ciudad,
        df_traslados=df_traslados,
        df_custodia=df_custodia,
        top_n=top_n,
        modo="estricto"
    )

    resultado_estimado = optimizar_rotacion(
        placas=placas_inventario,
        df_inventario=df_inventario_rotar_final,
        df_stats=df_referencia_ciudad,
        df_traslados=df_traslados,
        df_custodia=df_custodia,
        top_n=top_n,
        modo="estimado"
    )

    pack_estricto = armar_paquete_v4(
        resultado_estricto,
        df_inventario_rotar_final,
        df_referencia_ciudad,
        df_traslados,
        "estricto",
        map_fecha,
        map_dias_ingreso
    )

    pack_estimado = armar_paquete_v4(
        resultado_estimado,
        df_inventario_rotar_final,
        df_referencia_ciudad,
        df_traslados,
        "estimado",
        map_fecha,
        map_dias_ingreso
    )

    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:

        # TRASLADOS
        pack_estricto["traslados"].to_excel(
            writer, sheet_name="Traslados", index=False
        )

        # ESCENARIOS ESTRICTO
        pack_estricto["resultado"].to_excel(
            writer, sheet_name="Escenarios", index=False
        )

        pack_estricto["resumen"].to_excel(
            writer, sheet_name="Resumen", index=False
        )

        # ESTIMADO
        pack_estimado["resultado"].to_excel(
            writer, sheet_name="Escenarios_Estimado", index=False
        )

        pack_estimado["no_movibles"].to_excel(
            writer, sheet_name="No_Opt_Estimado", index=False
        )

    print(f"✅ Listo: '{filename}'")

    return filename


# =====================================================
# FORMATEADOR
# =====================================================

def formatear_excel_profesional(
    input_file: str,
    output_file: str = "Reporte_Modelo_Formateado.xlsx"
):

    excel_file = pd.ExcelFile(input_file)

    with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column

        # =========================
        # TABLA
        # =========================
        if max_row > 1:
            table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
            table = Table(
                displayName=f"Tabla_{sheet_name[:15]}",
                ref=table_ref
            )

            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            table.tableStyleInfo = style
            ws.add_table(table)

        # =========================
        # ANCHO COLUMNAS
        # =========================
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        # =========================
        # FORMATOS
        # =========================
        for col in range(1, max_col + 1):
            header = ws.cell(row=1, column=col).value

            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=col)

                if "Fecha" in str(header):
                    cell.number_format = FORMAT_DATE_XLSX14

                if any(
                    palabra in str(header)
                    for palabra in [
                        "Precio",
                        "Costo",
                        "Margen",
                        "Diferencia"
                    ]
                ):
                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

                if header in [
                    "Rank",
                    "Rank_Global",
                    "Dias_Esperados_Origen",
                    "Dias_Esperados_Destino"
                ]:
                    cell.alignment = Alignment(horizontal="center")

    wb.save(output_file)

    print(f"✅ Archivo formateado listo: {output_file}")